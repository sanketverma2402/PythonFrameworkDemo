import openpyxl

workbook = openpyxl.load_workbook("D:\\Other_Notes\\Python\\Workspace\\1stFebPythonSelenium_Framework\\TestData\Data.xlsx")
sheet=workbook['Sheet1']     #workbook.active

#get row Size in a sheet
rowSize=sheet.max_row
print(rowSize)

#get col size
colSize=sheet.max_column
print(colSize)

#get data from sheet1
s1=sheet["A1"].value
print(s1)

print(sheet["A1"].value)

print("--------")
#alternate
s2=sheet.cell(row=1,column=1).value
print(s2)

print(sheet.cell(row=1,column=1).value)



